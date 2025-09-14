import os
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


# Step 1: Locate the latest Ashok Leyland PDF
download_dir = "ashok_leyland_downloads"  # folder where PDFs are stored
pdf_files = [os.path.join(download_dir, f) for f in os.listdir(download_dir) if f.endswith(".pdf")]

if not pdf_files:
    raise FileNotFoundError("No PDF files found in ashok_leyland_downloads")

latest_pdf = max(pdf_files, key=os.path.getctime)  # pick the most recent file
print(f"Using latest PDF: {latest_pdf}")

excel_file = os.path.join(download_dir, "ashok_leyland_domestic.xlsx")


# Step 2: Extract "Domestic" sales table
domestic_rows = []

with pdfplumber.open(latest_pdf) as pdf:
    for page in pdf.pages:
        tables = page.extract_tables({
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
            "intersection_tolerance": 5
        })

        for table in tables:
            for row in table:
                if not any(str(c).strip() for c in row):
                    continue

                first_cell = str(row[0]).strip().upper()

                # start capturing rows from "M&HCV Trucks"
                if "M&HCV TRUCKS" in first_cell or domestic_rows:
                    domestic_rows.append(row)

                    # stop at "Total Vehicles"
                    if "TOTAL VEHICLES" in first_cell:
                        break

            if domestic_rows:
                break
        if domestic_rows:
            break

if not domestic_rows:
    raise ValueError("Could not find Domestic block in PDF")


# Step 3: Clean data into DataFrame
df = pd.DataFrame(domestic_rows)
df = df.dropna(axis=1, how="all")       # drop empty columns
df = df.iloc[:, :7]                     # keep only first 7 cols

df.columns = [
    "CATEGORY",
    "May'25",
    "May'24",
    "Inc/Dec (Month)",
    "May'25 (Cumulative)",
    "May'24 (Cumulative)",
    "Inc/Dec (Cumulative)"
]


# Step 4: Save to Excel and format
df.to_excel(excel_file, index=False)

wb = load_workbook(excel_file)
ws = wb.active

# format header
for cell in ws[1]:
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")

# align data cells
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# adjust column widths
for col in ws.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws.column_dimensions[col[0].column_letter].width = max_length + 2

wb.save(excel_file)
print(f"Domestic sales data saved: {excel_file}")


# Step 5: Print output neatly
print("\n--- Ashok Leyland Domestic Sales Data ---\n")
print(df.to_string(index=False))
